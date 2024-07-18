from pathlib import Path
import os
from openpyxl import load_workbook
import csv
import numpy as np
import warnings
import copy
from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter.simpledialog import askstring

warnings.simplefilter(action='ignore', category=FutureWarning)

mapaDeUnidadesOrcamentarias = {
    "117;122" : "1489",
    "117;366" : "1451",
    "117;361" : "1154",
    "117;365" : "1159",
    "101;122" : "1397",
    "101;128" : "1397",
    "101;361" : "7",
    "101;362" : "189",
    "101;364" : "307",
    "101;365" : "712",
    "101;366" : "1415",
    "101;367" : "1416",
    "101;368" : "7",
    "101;812" : "602",
    "100;122" : "1505",
    "100;361" : "1373",
    "100;365" : "1382",
    "100;366" : "1469",
    "100;367" : "1470",
    "113;361" : "771",
    "113;365" : "808",
    "113;366" : "1475",
    "114;361" : "1147",
    "115;361" : "867",
    "115;362" : "1149",
    "115;367" : "1434",
    "115;365" : "904",
    "115;366" : "1433",
    "116;361" : "933",
    "116;365" : "969",
    "119;122" : "1499",
    "119;361" : "1176",
    "119;362" : "1190",
    "119;365" : "1229",
    "119;812" : "602",
    "108;361" : "1346",
    "108;365" : "1350",
    "108;366" : "1425",
    "108;367" : "1426",
    "106;361" : "32",
    "106;365" : "730",
    "106;366" : "1421",
    "106;367" : "1422",
    "107;361" : "1670",
    "107;365" : "1677",
    "107;366" : "1671",
    "107;367" : "1672",
    "109;361" : "1689",
    "109;365" : "1696",
    "109;366" : "1690",
    "109;367" : "1691",
    "104;361" : "1519",
    "104;365" : "1526",
    "104;366" : "1520",
    "104;367" : "1521",
    "105;361" : "1651",
    "105;365" : "1658",
    "105;366" : "1652",
    "105;367" : "1653",
    "118;122" : "1494",
    "118;361" : "1161",
    "118;365" : "1165",
    "118;366" : "1457",
    "118;812" : "602",
    "110;361" : "1757",
    "110;365" : "1762",
    "112;361" : "1249",
    "122;361" : "1578",
    "122;365" : "1625",
    "164;122" : "1397",
    
}

arquivoGeradoDespesas = []
despesasGeradasOficial = []
arquivoGeradoReceitas = []
dataReceitas = []
dataDespesas = []
despesasFiltradas = []
arquivoSiopeZerado = []
listaDeSiopeArquivoZerado = []
listaDeArquivosGerados = []
listaDeReceitasNaPasta = []
arrayTesteDesp = []
logGeracao = []
siopeArquivoZerado = 'SIOPE_ZERADO.csv'
receitasNaPasta = 'receitas'
despesasNaPasta = 'despesas'
script_dir = os.path.dirname(__file__)
global despesaLocalizadaSiope

def varClear():
    arquivoGeradoDespesas.clear()
    despesasGeradasOficial.clear()
    arquivoGeradoReceitas.clear()
    dataReceitas.clear()
    dataDespesas.clear()
    despesasFiltradas.clear()
    arquivoSiopeZerado.clear()
    listaDeSiopeArquivoZerado.clear()
    listaDeArquivosGerados.clear()
    listaDeReceitasNaPasta.clear()
    arrayTesteDesp.clear()

def selecionarArquivos():
    botaoGerar["state"] = "disabled"
    janela.update()
    varClear()
    global caminho
    global arquivoSiopeZerado
    base = askdirectory()
    script_dir = os.path.dirname(__file__) #<-- absolute dir the script is in
    rel_path = base + "/"
    caminho = os.path.join(script_dir, rel_path)
    pathLote = Path(caminho)
    
    with open(script_dir + "/SIOPE_ZERADO.csv", mode='r', newline='') as arquivoSiope:
        leitor = csv.reader(arquivoSiope, delimiter=";")
        arquivoSiopeZerado = list(leitor)
            
    #Receitas
    workbook = load_workbook(filename = caminho + 'receitas.xlsx')

    sheet = workbook.active

    for row in sheet.iter_rows():
        arrayRowReceita = []
        for cell in row:
            arrayRowReceita.append(cell.value if str(cell.value).split(",")[0] != "" else 0)
        dataReceitas.append(arrayRowReceita)
    
    #Despesas
    workbook = load_workbook(filename = caminho + 'despesas.xlsx')

    sheet = workbook.active

    for row in sheet.iter_rows():
        arrayRowDespesa = []
        for cell in row:
            # if cell.value is not None and str(cell.value).split(",")[0] != "TOTAL" and str(cell.value).split(",")[0] != "" and str(cell.value).split(",")[0] != "TOTAL GERAL":
            arrayRowDespesa.append((str(cell.value).split(",") if str(cell.value).split(",")[0] != "" else ['0']))
                
        dataDespesas.append(arrayRowDespesa)
    
    botaoGerar["state"] = "normal"
    
def remove_repetidos(listaComDuplicados):
    l = []
    for i in listaComDuplicados:
        if i not in l:
            l.append(i)
        else:
            continue
    return l 
        
def gerarDespesas():
    numDespNaoLocalizadas = 0
    for i, despesa in enumerate(dataDespesas, start = 0):
        for key in mapaDeUnidadesOrcamentarias:
            findUnidadeOrc = False
            if len(despesa) > 13:
                if key == str(despesa[7][0])+";"+str(despesa[4][0]):
                    del(dataDespesas[i][8])
                    dataDespesas[i].append(mapaDeUnidadesOrcamentarias[key])
                    findUnidadeOrc = True
                    break
                
            elif len(despesa) == 13:  
                if key == despesa[7][0]+";"+str(despesa[4][0]):
                    dataDespesas[i].append(mapaDeUnidadesOrcamentarias[key])
                    findUnidadeOrc = True
                    break
                
        if findUnidadeOrc == False:
            numDespNaoLocalizadas = numDespNaoLocalizadas + 1
            askstring("Despesa", "Despesa não localizada! " + str(despesa))
            logGeracao.append("Despesa não localizada! " + "Despesa: " + str(despesa[6][0]) 
                            + " Fonte Recurso:" + str(despesa[7][0]) + " SubFunção: " + str(despesa[4][0]))
            
    if numDespNaoLocalizadas > 0:
        return 0
        
    global despesaLocalizadaSiope
    for despesa in dataDespesas:
        despesaLocalizadaSiope = False
        digitado = False
        while despesaLocalizadaSiope == False:
            for x, linhaSiopZerado in enumerate (arquivoSiopeZerado, start=1):
                # print(despesa)
                if len(linhaSiopZerado) < 9:
                    continue
                
                elif despesa[6][0] == '':
                    despesa[6][0] = '" "'
                    break
                
                elif len(despesa[6][0]) < 8:
                    break
                
                elif digitado and ";" + str(despesa[13]) + ";" + str((despesa[6][0]).replace(".","")[0:8]) == ";" + str(linhaSiopZerado[2] + ";" + linhaSiopZerado[3]):
                    arquivoGeradoDespesas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3])
                                        + ";" + linhaSiopZerado[4] + ";"
                                        + str(despesa[9][0]).replace(".",",") + ";" + str(despesa[10][0]).replace(".",",") + ";" + str(despesa[11][0]).replace(".",",")
                                        + ";" + str(despesa[12][0] + ";" + "DIG").replace(".",",")).split(";"))
                    despesaLocalizadaSiope = True
                    break
                
                elif ";" + str(despesa[13]) + ";" + str((despesa[6][0]).replace(".","")[0:8]) == ";" + str(linhaSiopZerado[2] + ";" + linhaSiopZerado[3]):
                    arquivoGeradoDespesas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3])
                                        + ";" + linhaSiopZerado[4] + ";"
                                        + str(despesa[9][0]).replace(".",",") + ";" + str(despesa[10][0]).replace(".",",") + ";" + str(despesa[11][0]).replace(".",",")
                                        + ";" + str(despesa[12][0]).replace(".",",")).split(";"))
                    despesaLocalizadaSiope = True
                    printarNaTela(despesa)
                    janela.update()
                    break
                
            if despesaLocalizadaSiope == False:
                printarNaTela("Despesa " + despesa[6][0] + " não localizada na tabela SIOPE. Digite um elemento existente.")
                despesaDigitada = askstring("Despesa", "Despesa " + despesa[6][0] + " não localizada na tabela SIOPE. Digite um elemento existente:")
                despesa[6][0] = despesaDigitada
                digitado == True
                
                
    for i, elementoDespesa in enumerate(arquivoGeradoDespesas, start=0):
        if str(arquivoGeradoDespesas[i][3][0:6]) == "319004":
            arquivoGeradoDespesas[i][3] = arquivoGeradoDespesas[i][3][0:6] + "01"
            continue
        
        elif str(arquivoGeradoDespesas[i][3][0:6]) == "319011":
            arquivoGeradoDespesas[i][3] = arquivoGeradoDespesas[i][3][0:6] + "01"
            continue
        
        elif str(arquivoGeradoDespesas[i][3][0:6]) == "319013":
            arquivoGeradoDespesas[i][3] = arquivoGeradoDespesas[i][3][0:6] + "02"
            continue
        
        elif str(arquivoGeradoDespesas[i][3][0:6]) == "319113":
            arquivoGeradoDespesas[i][3] = arquivoGeradoDespesas[i][3][0:6] + "08"
            continue
        
        elif len(arquivoGeradoDespesas[i]) > 9:
            del arquivoGeradoDespesas[i][9]
            continue
            
        arquivoGeradoDespesas[i][3] = arquivoGeradoDespesas[i][3][0:6] + "99"
        
def ajustarDespesas():
    arquivoGeradoDespesasCopia = copy.deepcopy(arquivoGeradoDespesas)
    for x, aux in enumerate(arquivoGeradoDespesasCopia, start=0):
        arquivoGeradoDespesasCopia[x][5] = '0'
        arquivoGeradoDespesasCopia[x][6] = '0'
        arquivoGeradoDespesasCopia[x][7] = '0'
        arquivoGeradoDespesasCopia[x][8] = '0'
        
    arquivoGeradoDespesasCopia = remove_repetidos(arquivoGeradoDespesasCopia)
    despesasGeradasOficial = copy.deepcopy(arquivoGeradoDespesasCopia)
    
    for i, despesaCopia in enumerate(arquivoGeradoDespesasCopia, start=0):
        for despesa in arquivoGeradoDespesas:
            if str(despesaCopia[2] + despesaCopia[3] + despesaCopia[4]) == str(despesa[2] + despesa[3] + despesa[4]):
                auxDespesaOficial = float(despesasGeradasOficial[i][5].replace(",",".") if despesasGeradasOficial[i][5] != '' else 0)
                auxDespesaGerada = float(despesa[5].replace(",",".") if arquivoGeradoDespesas[i][5] != '' else 0)
                despesasGeradasOficial[i][5] = str("%.2f" % (auxDespesaGerada + auxDespesaOficial)).replace(".",",")
                
                auxDespesaOficial = float(despesasGeradasOficial[i][6].replace(",",".") if despesasGeradasOficial[i][6] != '' else 0)
                auxDespesaGerada = float(despesa[6].replace(",",".") if arquivoGeradoDespesas[i][6] != '' else 0)
                despesasGeradasOficial[i][6] = str("%.2f" % (auxDespesaGerada + auxDespesaOficial)).replace(".",",")
                
                auxDespesaOficial = float(despesasGeradasOficial[i][7].replace(",",".") if despesasGeradasOficial[i][7] != '' else 0)
                auxDespesaGerada = float(despesa[7].replace(",",".") if arquivoGeradoDespesas[i][7] != '' else 0)
                despesasGeradasOficial[i][7] = str("%.2f" % (auxDespesaGerada + auxDespesaOficial)).replace(".",",")
                
                auxDespesaOficial = float(despesasGeradasOficial[i][8].replace(",",".") if despesasGeradasOficial[i][8] != '' else 0)
                auxDespesaGerada = float(despesa[8].replace(",",".") if arquivoGeradoDespesas[i][8] != '' else 0)
                despesasGeradasOficial[i][8] = str("%.2f" % (auxDespesaGerada + auxDespesaOficial)).replace(".",",")
                
    for n, despGOficial in enumerate(despesasGeradasOficial, start=0):
        localizado = False
        for siope in arquivoSiopeZerado:
            if len(siope) < 9:
                    continue
                
            elif str(siope[2] + ";" + siope[3]) in str(despGOficial[2] + ";" + despGOficial[3]):
                localizado = True
                listaDeArquivosGerados.append(';'.join(despGOficial))
                printarNaTela(despGOficial)
                janela.update()
                break
            
        if localizado == False:
            despGOficial[3] = despGOficial[3][0:6] + "00"
            listaDeArquivosGerados.append(';'.join(despGOficial))
            printarNaTela(despGOficial)
            janela.update()
            
def gerarReceitas():
    for x, linhaReceita in enumerate(dataReceitas, start=0):
        printarNaTela(linhaReceita)
        janela.update()
        for y, linhaSiopZerado in enumerate (arquivoSiopeZerado, start=1):
            # if str(linhaReceita[0]).replace(".","")[0:8] in str(linhaSiopZerado):
            if str(linhaReceita[0]).replace(".","")[0:8] in str(linhaSiopZerado) or str(linhaReceita[0]).replace(".","")[2:8] in str(linhaSiopZerado):
                # if str(linhaReceita[0]).replace(".","")[:1] == "1" :
                arrAuxReceita = np.array(arquivoGeradoReceitas)
                i = np.argwhere(arrAuxReceita == str(linhaSiopZerado[3]))
                
                if str(linhaReceita[0]).replace(".","")[:2] != "95" and str(linhaReceita[0]).replace(".","")[:2] != "99" and str(linhaReceita[0]).replace(".","")[:1] != "7":
                
                    if len(i) > 0:
                        valorInicial = float(arquivoGeradoReceitas[i[0][0]][5].replace(",",".") if arquivoGeradoReceitas[i[0][0]][5] != '' else 0)
                        valorInicialDuplicado = float(linhaReceita[5] if linhaReceita[5] != '' else 0)
                        arquivoGeradoReceitas[i[0][0]][5] = str("%.2f" % (valorInicial + valorInicialDuplicado)).replace(".",",")
                        
                        valorArrecadacaoAteOMes = float(arquivoGeradoReceitas[i[0][0]][6].replace(",",".") if arquivoGeradoReceitas[i[0][0]][6] != '' else 0)
                        valorArrecadacaoAteOMesDuplicado = float(linhaReceita[4] if linhaReceita[4] != '' else 0)
                        arquivoGeradoReceitas[i[0][0]][6] = str("%.2f" % (valorArrecadacaoAteOMes + valorArrecadacaoAteOMesDuplicado)).replace(".",",")
                        
                    else:
                        # print((linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4])
                        arquivoGeradoReceitas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3])
                                        + ";" + linhaSiopZerado[4] + ";"
                                        + str(linhaReceita[5]).replace(".",",") + ";" + str(linhaReceita[4]).replace(".",",") + ";"
                                        + ";" + ";").split(";"))
                        
                #IntraOrçamentaria       
                elif str(linhaReceita[0]).replace(".","")[:1] == "7":
                    # print("IntraOrçamentaria " + linhaReceita[0].replace(".",""))
                    arrAux = np.array(arquivoGeradoReceitas)
                    # print(linhaSiopZerado[3])
                    # print(linhaSiopZerado[3][:5])
                    
                    i = np.argwhere(np.core.defchararray.find(arquivoGeradoReceitas, linhaSiopZerado[3][:5])!=-1)
                    if len(i) > 0:
                        # print(i)
                        
                        linhaArquivoGerado = arquivoGeradoReceitas[i[0][0]]
                        varAux = str("%.2f" % (float(str(linhaReceita[4]).replace("-","") if linhaReceita[4] != '' else 0) + float(linhaArquivoGerado[9].replace(",",".") if linhaArquivoGerado[9] != '' else 0))).replace(".",",")
                        linhaArquivoGerado[9] = varAux
                    else:
                        # print((linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4])
                        arquivoGeradoReceitas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4] + ";"
                                        + ";" + ";" + ";" + ";" + str(linhaReceita[4]).replace(".",",")).split(";"))
                    # print(arquivoGeradoReceitas[i[0][0]])
                    # print(i)
                
                #Deduções fundeb    
                elif str(linhaReceita[0]).replace(".","")[:2] == "95":
                    # print("Deduções fundeb " + linhaReceita[0].replace(".",""))
                    arrAux = np.array(arquivoGeradoReceitas)
                    # print(linhaSiopZerado[3])
                    # print(linhaSiopZerado[3][:5])
                    
                    i = np.argwhere(np.core.defchararray.find(arquivoGeradoReceitas, linhaSiopZerado[3][:5])!=-1)
                    # print(arquivoGeradoReceitas[i[0][0]])
                    if len(i) > 0:
                        # print(i)
                        
                        linhaArquivoGerado = arquivoGeradoReceitas[i[0][0]]
                        varAux = str("%.2f" % (float(str(linhaReceita[4]).replace("-","") if linhaReceita[4] != '' else 0) + float(linhaArquivoGerado[7].replace(",",".") if linhaArquivoGerado[7] != '' else 0))).replace(".",",")
                        linhaArquivoGerado[7] = varAux
                    else:
                        # print((linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4])
                        arquivoGeradoReceitas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4] + ";"
                                        + ";" + ";" + str(linhaReceita[4]).replace(".",",") + ";" +";").split(";"))
                    
                    # print(i)
                    
                #Outras deduções
                elif str(linhaReceita[0]).replace(".","")[:2] == "99":
                    
                    # print("Deduções fundeb " + linhaReceita[0].replace(".",""))
                    
                    # print(linhaSiopZerado[3])
                    # print(linhaSiopZerado[3][:5])
                    
                    arrAux = np.array(arquivoGeradoReceitas)
                    i = np.argwhere(np.core.defchararray.find(arquivoGeradoReceitas, linhaSiopZerado[3][:5])!=-1)
                    # print(i)
                    
                    if len(i) > 0:
                        # print(i)
                        
                        linhaArquivoGerado = arquivoGeradoReceitas[i[0][0]]
                        varAux = str("%.2f" % (float(str(linhaReceita[4]).replace("-","") if linhaReceita[4] != '' else 0) + float(linhaArquivoGerado[8].replace(",",".") if linhaArquivoGerado[8] != '' else 0))).replace(".",",")
                        linhaArquivoGerado[8] = varAux
                        # print(arquivoGeradoReceitas[i[0][0]])
                        # print(i)
                    else:
                        # print((linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4])
                        arquivoGeradoReceitas.append((linhaSiopZerado[0] + ";" + linhaSiopZerado[1] + ";" + linhaSiopZerado[2] + ";" 
                                        + (linhaSiopZerado[3] if linhaSiopZerado[3][-2:] != "00" else linhaSiopZerado[3][0:6] + "01")  + ";" + linhaSiopZerado[4] + ";"
                                        + ";" + ";"
                                        + ";"  + str(linhaReceita[4]).replace(".",",") +";").split(";"))
                
                break
        
            elif len(str(linhaReceita[0]).replace(".","")[:-4]) < 8:
                break
            
    for geradoReceita in arquivoGeradoReceitas:
        listaDeArquivosGerados.append(';'.join(geradoReceita))
        printarNaTela(geradoReceita)
        janela.update()

def escreverArquivo():
    if logGeracao:
        geracaoDoArquivo = Path(caminho + "Log.txt")
        geracaoDoArquivo.write_text("\n".join(logGeracao))
        printarNaTela("Verifique o log da geração!")
    else:
        geracaoDoArquivo = Path(caminho + "ArquivoGerado.csv")
        geracaoDoArquivo.write_text("\n".join(listaDeArquivosGerados))
        printarNaTela("Geração concluida!")

def printarNaTela(msg):
    global label
    label["text"] = msg

def mainFunc():
    if not dataDespesas or not dataReceitas:
        printarNaTela("Selecione o local dos arquivos!")
        janela.update()
        
    else:
        if listaDeArquivosGerados:
            varClear()
            
        botaoGerar["state"] = "disabled"
        gerarReceitas()
        gerarDespesas()
        ajustarDespesas()
        escreverArquivo()
        botaoGerar["state"] = "normal"

def CenterWindowToDisplay(Screen: Tk, width: int, height: int):
    """Centers the window to the main display/monitor"""
    screen_width = Screen.winfo_screenwidth()
    screen_height = Screen.winfo_screenheight()
    x = int((screen_width/2) - (width/2))
    y = int((screen_height/2) - (height/1.5))
    return f"{width}x{height}+{x}+{y}"

janela = Tk()
janela.geometry(CenterWindowToDisplay(janela, 550, 250)) 
janela.resizable(0,0)
janela.title("Gerar SIOPE")

textoOrientacao = Label(janela, wraplength = 240, text="Selecione o local dos arquivos: ", font=("Arial", 12))
textoOrientacao.place(x=150, y=50)

# m = StringVar() 
# boxSistemas = ttk.Combobox(janela, textvariable = m)
# boxSistemas['values'] = (sistemas)
# boxSistemas.place(x=45, y=70)
icon = PhotoImage(file= script_dir + "/img/pasta.png")
buscarArquivos = Button(janela, image=icon, command=selecionarArquivos)
buscarArquivos.place(x=370, y=50)


botaoGerar = Button(janela, text="Gerar", command=mainFunc, width=10)
botaoGerar.place(x=230, y=95)

label = Label(janela, wraplength=530, font=("Arial", 11))
label.place(x=30, y=145)

janela.mainloop()
# source .venv/Scripts/activate